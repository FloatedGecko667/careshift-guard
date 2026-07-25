"""
勤務形態一覧表（従業者の勤務の体制及び勤務形態一覧表）の Excel 出力。

厚生労働省の標準様式が求める必須10項目
  1 サービス種別 / 2 事業所名 / 3 職種 / 4 勤務形態（常勤・非常勤／専従・兼務）
  5 資格・修了研修 / 6 氏名 / 7 従業者ごとの日々の勤務時間
  8 従業者ごとの当月の勤務時間数合計 / 9 兼務状況 / 10 常勤の従業者が勤務すべき時間数
に加え、通所介護で追加が求められる
  利用者数 / サービス提供の単位 / サービス提供時間帯 / 勤務延時間数
を出力する。

実装上の要点
  ・合計と常勤換算は「値」ではなく Excel の数式で書き込む。
    運営指導の場で担当者がセルを開いて検算できることに意味がある。
  ・常勤換算後の人数は小数点以下第2位を切り捨てる（ROUNDDOWN(x, 1)）。
  ・管理者（manager）は常勤換算の対象外。行としては出力するが、
    基準判定のサマリには算入しない。
  ・兼務者は職種ごとに行を分け、勤務時間を従事割合で按分する。

License: MIT
"""
from __future__ import annotations

import datetime as dt

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.jobs import FTE_TARGET_JOBS, JOB_ORDER
from app.solver import REST, Problem, Solution

# ---------------------------------------------------------------- 表示定義
# 職種の表示順と判定対象は app/jobs.py に集約している。
# ここで再定義すると表記が食い違うため、再輸出のみ行う。
__all__ = ["FTE_TARGET_JOBS", "JOB_ORDER", "build_workbook", "expand_rows", "export"]

WD = ["月", "火", "水", "木", "金", "土", "日"]
WORK_FORM_LABEL = {
    "A": "A 常勤・専従", "B": "B 常勤・兼務",
    "C": "C 常勤以外・専従", "D": "D 常勤以外・兼務",
}

THIN = Side(style="thin", color="9AA3AD")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_HEAD = PatternFill("solid", fgColor="DEEAF6")
F_SUB = PatternFill("solid", fgColor="F2F5F9")
F_SAT = PatternFill("solid", fgColor="EEF4FB")
F_SUN = PatternFill("solid", fgColor="FBEEF0")
F_CLOSED = PatternFill("solid", fgColor="ECEFF3")
F_NG = PatternFill("solid", fgColor="FDECEA")
F_OK = PatternFill("solid", fgColor="EAF6EE")
FONT = "Meiryo UI"


def _c(ws, row, col, value=None, *, bold=False, size=9, fill=None,
       align="center", fmt=None, wrap=False):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = Font(name=FONT, size=size, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = BOX
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ---------------------------------------------------------------- 行の展開
def expand_rows(prob: Problem):
    """兼務者を職種ごとの行に分解する。戻り値は (職種, 職員index, 従事割合) の列。"""
    out = []
    for i, st in enumerate(prob.staff):
        primary = 1.0 - (st.secondary_ratio if st.secondary_job else 0.0)
        out.append((st.job, i, primary))
        if st.secondary_job:
            out.append((st.secondary_job, i, st.secondary_ratio))
    order = {j: k for k, j in enumerate(JOB_ORDER)}
    out.sort(key=lambda t: (order.get(t[0], 99), t[1]))
    return out


# ---------------------------------------------------------------- 本体
def build_workbook(prob: Problem, sol: Solution, *, year: int, month: int,
                   office_name: str, service_name: str = "通所介護",
                   avg_users: float = 22.0, service_start: str = "9:00",
                   service_end: str = "16:00", unit_name: str = "1単位") -> Workbook:
    first = dt.date(year, month, 1)
    nd = prob.num_days
    h = [p.work_minutes for p in prob.patterns]
    month_hours = prob.fulltime_month_minutes / 60
    week_hours = prob.fulltime_week_minutes / 60

    wb = Workbook()

    # =================================================================
    # シート1  勤務形態一覧表
    # =================================================================
    ws = wb.active
    ws.title = "勤務形態一覧表"
    ws.sheet_view.showGridLines = False

    DAY0 = 5                      # 日別列の開始（E列）
    COL_TOTAL = DAY0 + nd         # 当月合計
    COL_AVGW = COL_TOTAL + 1      # 週平均
    COL_NOTE = COL_TOTAL + 2      # 兼務状況
    LAST = COL_NOTE

    # ---- 印刷設定（A3横・幅を1ページに収める） ----
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 8            # A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "8:9"

    # ---- 表題と事業所情報 ----
    # 全列にまたがってマージすると、中央寄せの文字が可視範囲の外へ出てしまう。
    # 先頭の広い列だけにマージし、左寄せにする。
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t = ws.cell(row=1, column=1,
                value=f"従業者の勤務の体制及び勤務形態一覧表（{year}年{month}月分）")
    t.font = Font(name=FONT, size=13, bold=True)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    info = [
        ("サービス種類", service_name, "事業所名", office_name),
        ("利用者数（平均）", f"{avg_users:.0f} 名", "サービス提供の単位", unit_name),
        ("サービス提供時間帯", f"{service_start}〜{service_end}",
         "サービス提供時間数", f"{_svc_hours(service_start, service_end):.1f} 時間"),
        ("常勤の従業者が勤務すべき時間数",
         f"週 {week_hours:.1f} 時間 ／ 暦月 {month_hours:.1f} 時間",
         "対象期間", f"{first} 〜 {first + dt.timedelta(days=nd - 1)}"),
    ]
    for k, (l1, v1, l2, v2) in enumerate(info):
        r = 3 + k
        _c(ws, r, 1, l1, bold=True, fill=F_SUB, align="left", wrap=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        _c(ws, r, 2, v1, align="left")
        # 5列目以降は1列4.4幅の日付列なので、ラベルは複数列をマージして確保する
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        _c(ws, r, 4, l2, bold=True, fill=F_SUB, align="left")
        ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=17)
        _c(ws, r, 9, v2, align="left")
    ws.row_dimensions[6].height = 26

    # ---- 表ヘッダ（2段） ----
    HR = 8
    for col, label, width in ((1, "職種", 15), (2, "勤務形態", 15),
                              (3, "資格・修了研修", 20), (4, "氏名", 14)):
        ws.merge_cells(start_row=HR, start_column=col, end_row=HR + 1, end_column=col)
        _c(ws, HR, col, label, bold=True, fill=F_HEAD, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    closed = [prob.required_head["介護職員"][d] == 0 for d in range(nd)]
    for d in range(nd):
        date = first + dt.timedelta(days=d)
        col = DAY0 + d
        fill = F_CLOSED if closed[d] else (
            F_SUN if date.weekday() == 6 else (F_SAT if date.weekday() == 5 else F_HEAD))
        _c(ws, HR, col, date.day, bold=True, fill=fill)
        _c(ws, HR + 1, col, WD[date.weekday()], bold=True, fill=fill, size=8)
        ws.column_dimensions[get_column_letter(col)].width = 4.4

    for col, label, width in ((COL_TOTAL, "当月合計\n(時間)", 11),
                              (COL_AVGW, "週平均\n(時間)", 10),
                              (COL_NOTE, "兼務状況", 22)):
        ws.merge_cells(start_row=HR, start_column=col, end_row=HR + 1, end_column=col)
        _c(ws, HR, col, label, bold=True, fill=F_HEAD, wrap=True)
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = ws.cell(row=HR + 2, column=DAY0)

    # ---- 明細 ----
    rows = expand_rows(prob)
    weeks = nd / 7
    r = HR + 2
    job_rows: dict[str, list[int]] = {}
    for job, i, ratio in rows:
        st = prob.staff[i]
        job_rows.setdefault(job, []).append(r)

        _c(ws, r, 1, job, align="left")
        _c(ws, r, 2, WORK_FORM_LABEL[st.work_form_code], align="left", size=8)
        _c(ws, r, 3, "、".join(st.qualifications) if st.qualifications else "—",
           align="left", size=8, wrap=True)
        name = st.name if ratio == 1.0 else f"{st.name}（{ratio:.0%}）"
        _c(ws, r, 4, name, align="left")

        for d in range(nd):
            p = sol.assign[i, d]
            col = DAY0 + d
            v = round(h[p] * ratio / 60, 2) if p != REST else None
            fill = F_CLOSED if closed[d] else None
            _c(ws, r, col, v, fmt="0.0", fill=fill, size=8)

        a, b = get_column_letter(DAY0), get_column_letter(DAY0 + nd - 1)
        _c(ws, r, COL_TOTAL, f"=ROUND(SUM({a}{r}:{b}{r}),1)", bold=True, fmt="0.0")
        _c(ws, r, COL_AVGW,
           f"=ROUND({get_column_letter(COL_TOTAL)}{r}/{weeks:.4f},1)", fmt="0.0")
        note = ""
        if st.secondary_job:
            other = st.secondary_job if job == st.job else st.job
            note = f"{other} と兼務（当欄は{ratio:.0%}分）"
        _c(ws, r, COL_NOTE, note or "—", align="left", size=8, wrap=True)
        r += 1

    # ---- 職種別の常勤換算サマリ ----
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _c(ws, r, 1, "職種別 常勤換算後の人数と人員配置基準の判定",
       bold=True, size=11, align="left", fill=F_SUB)
    r += 1

    # 5列目以降は狭い日付列なので、サマリはマージして列幅を確保する
    SPANS = [(1, 1), (2, 2), (3, 3), (4, 4), (5, 8), (9, 12)]
    heads = ["職種", "勤務延時間数\n(当月・時間)", "常勤が勤務すべき\n時間数(時間)",
             "常勤換算後の人数", "基準上必要な数", "判定"]
    for (c0, c1), hd in zip(SPANS, heads, strict=True):
        if c1 > c0:
            ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
        _c(ws, r, c0, hd, bold=True, fill=F_HEAD, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1

    ng_days = {(v.job, v.day) for v in sol.violations}
    for job in FTE_TARGET_JOBS:
        rs = job_rows.get(job, [])
        tc = get_column_letter(COL_TOTAL)
        expr = "+".join(f"{tc}{x}" for x in rs) if rs else "0"
        req = max(prob.required_fte[job][d] for d in range(nd))
        bad = any((job, d) in ng_days for d in range(nd))
        vals = [job, f"=ROUND({expr},1)", month_hours,
                # 常勤換算は小数点以下第2位を切り捨てる
                f"=ROUNDDOWN(B{r}/C{r},1)", req,
                "日別に要確認" if bad else "適合"]
        for k, ((c0, c1), v) in enumerate(zip(SPANS, vals, strict=True)):
            if c1 > c0:
                ws.merge_cells(start_row=r, start_column=c0, end_row=r, end_column=c1)
            _c(ws, r, c0, v,
               bold=(k in (0, 3, 5)),
               align="left" if k == 0 else "center",
               fmt="0.0" if k in (1, 2, 3, 4) else None,
               fill=(F_NG if bad else F_OK) if k == 5 else None)
        r += 1

    r += 1
    for line in (
        "※ 常勤換算後の人数 ＝ 勤務延時間数 ÷ 常勤の従業者が勤務すべき時間数"
        "（小数点以下第2位を切り捨て）",
        "※ 管理者は常勤換算の対象外のため、上表の判定には算入していない。",
        "※ 兼務者は職種ごとに行を分け、勤務時間を従事割合で按分して計上している。",
        "※ 上表は暦月を単位とした集計である。日ごとの人員配置基準の充足状況は"
        "「日別の基準判定」シートを参照。",
    ):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST)
        c = ws.cell(row=r, column=1, value=line)
        c.font = Font(name=FONT, size=8)
        c.alignment = Alignment(horizontal="left", vertical="center")
        r += 1

    # =================================================================
    # シート2  日別の基準判定
    # =================================================================
    ws2 = wb.create_sheet("日別の基準判定")
    ws2.sheet_view.showGridLines = False
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.paperSize = 8            # A3
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.print_title_rows = "3:3"
    # マージすると幅が足りず末尾が切れるため、マージせず隣の空セルへ溢れさせる
    t2 = ws2.cell(row=1, column=1,
                  value=f"日別の人員配置基準 充足状況（{year}年{month}月）")
    t2.font = Font(name=FONT, size=12, bold=True)
    t2.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[1].height = 22

    _c(ws2, 3, 1, "職種", bold=True, fill=F_HEAD)
    _c(ws2, 3, 2, "区分", bold=True, fill=F_HEAD)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12
    for d in range(nd):
        date = first + dt.timedelta(days=d)
        fill = F_CLOSED if closed[d] else (
            F_SUN if date.weekday() == 6 else (F_SAT if date.weekday() == 5 else F_HEAD))
        _c(ws2, 3, 3 + d, f"{date.day}\n{WD[date.weekday()]}",
           bold=True, fill=fill, size=8, wrap=True)
        ws2.column_dimensions[get_column_letter(3 + d)].width = 5.0
    ws2.freeze_panes = ws2.cell(row=4, column=3)

    rr = 4
    for job in FTE_TARGET_JOBS:
        for kind in ("必要", "実際"):
            _c(ws2, rr, 1, job if kind == "必要" else "", align="left",
               bold=(kind == "必要"))
            _c(ws2, rr, 2, kind, fill=F_SUB, size=8)
            for d in range(nd):
                val = prob.required_fte[job][d] if kind == "必要" else sol.fte[job, d]
                bad = kind == "実際" and (job, d) in ng_days
                fill = F_CLOSED if closed[d] else (F_NG if bad else None)
                _c(ws2, rr, 3 + d, round(val, 2), fmt="0.0", fill=fill, size=8,
                   bold=bad)
            rr += 1

    rr += 1
    _c(ws2, rr, 1, "判定", bold=True, fill=F_SUB, align="left")
    _c(ws2, rr, 2, "", fill=F_SUB)
    for d in range(nd):
        bad = any((j, d) in ng_days for j in FTE_TARGET_JOBS)
        _c(ws2, rr, 3 + d, "×" if bad else ("—" if closed[d] else "○"),
           bold=True, fill=F_CLOSED if closed[d] else (F_NG if bad else F_OK), size=9)

    rr += 2
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=12)
    c = ws2.cell(row=rr, column=1, value=(
        f"検出された基準違反：{len(sol.violations)} 件"
        "　／　×の日は人員基準欠如減算（基本報酬の30％減算）の対象になり得る。"))
    c.font = Font(name=FONT, size=9, bold=True,
                  color="C0392B" if sol.violations else "1E7E46")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # =================================================================
    # シート3  算定根拠
    # =================================================================
    ws3 = wb.create_sheet("算定根拠")
    ws3.sheet_view.showGridLines = False
    lines = [
        ("常勤換算の算定方法", True),
        ("", False),
        ("常勤換算後の人数 ＝ 当該職種の勤務延時間数 ÷ 常勤の従業者が勤務すべき時間数", False),
        (f"　　　　　　　　　＝ 勤務延時間数 ÷ {month_hours:.1f} 時間（暦月）", False),
        ("　　　　　　　　　　小数点以下第2位を切り捨てる", False),
        ("", False),
        ("人員配置基準（通所介護）", True),
        ("", False),
        ("　介護職員　　　　： 利用者15人までは常勤換算1以上。", False),
        ("　　　　　　　　　　15人を超える場合は、15を超える部分の数を5で除して得た数に", False),
        ("　　　　　　　　　　1を加えた数以上。除して得た数の端数は切り上げない。", False),
        (f"　　　　　　　　　　（利用者{avg_users:.0f}名 → 必要常勤換算 "
         f"{max(prob.required_fte['介護職員']):.1f}）", False),
        ("　看護職員　　　　： サービス提供日ごとに専従1以上。", False),
        ("　生活相談員　　　： サービス提供時間帯に応じ専従1以上。", False),
        ("　機能訓練指導員　： 専従1以上。", False),
        ("", False),
        ("常勤・非常勤の区分", True),
        ("", False),
        (f"　常勤　： 事業所が定める常勤職員の勤務すべき時間数"
         f"（週{week_hours:.1f}時間）に達している者。", False),
        ("　　　　　週32時間を下回る場合は32時間を基準とする。", False),
        ("　非常勤： 上記に達していない者。", False),
        ("", False),
        ("勤務形態の区分", True),
        ("", False),
        ("　A：常勤で専従　　B：常勤で兼務　　C：常勤以外で専従　　D：常勤以外で兼務", False),
        ("", False),
        ("留意事項", True),
        ("", False),
        ("　・管理者は常勤換算の対象外である。", False),
        ("　・兼務者は職種ごとに行を分け、勤務時間を従事割合で按分する。", False),
        ("　・常勤職員の有給休暇、勤務命令による研修参加、病欠は勤務時間に含めてよい。", False),
        ("　　ただし欠勤が暦月で1か月分を超える場合は常勤換算に含めない。", False),
        ("", False),
        ("このシートは CareShift Guard が自動生成した。", False),
        (f"シフト表の求解ステータス：{sol.status}／求解時間 {sol.solve_seconds:.2f} 秒", False),
    ]
    ws3.column_dimensions["A"].width = 110
    for k, (text, bold) in enumerate(lines):
        c = ws3.cell(row=2 + k, column=1, value=text)
        c.font = Font(name=FONT, size=11 if bold else 9.5, bold=bold)
        c.alignment = Alignment(horizontal="left", vertical="center")

    return wb


def _svc_hours(start: str, end: str) -> float:
    def m(s):
        hh, mm = s.split(":")
        return int(hh) * 60 + int(mm)
    return (m(end) - m(start)) / 60


def export(prob: Problem, sol: Solution, path: str, **kw) -> str:
    build_workbook(prob, sol, **kw).save(path)
    return path
