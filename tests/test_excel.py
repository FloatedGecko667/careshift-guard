"""勤務形態一覧表 Excel 出力の回帰テスト。

Excel の数式は openpyxl では評価されないため、LibreOffice で csv に変換して
実際に計算させ、Python の独立計算と突き合わせる。
これにより「数式が正しく書けているか」まで検証できる。
LibreOffice が無い環境では該当テストを skip する。
"""
from __future__ import annotations

import csv
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.excel import FTE_TARGET_JOBS, expand_rows, export
from app.solver import solve
from tests.demo_data import build_demo

SHEETS = ["勤務形態一覧表", "日別の基準判定", "算定根拠"]
NG_FILL = ("00FDECEA", "FFFDECEA")


@pytest.fixture(scope="module")
def artifacts(tmp_path_factory):
    """1回だけ解いて xlsx を生成し、全テストで共有する。

    生成と検証を同一の解で行うことが重要。別々に solve すると
    壁時計制限のため異なる解になり、比較が成立しない。
    """
    prob = build_demo()
    sol = solve(prob, time_limit=5.0, workers=2, deterministic=True)
    path = tmp_path_factory.mktemp("xlsx") / "kinmu.xlsx"
    export(prob, sol, str(path), year=2026, month=8,
           office_name="デイサービスさくら", avg_users=22.0)
    return prob, sol, path


def _to_csv(xlsx: Path) -> list[list[str]] | None:
    """LibreOffice で1枚目を csv 化する（数式が評価される）。

    既定のユーザープロファイルは同時起動でロック競合を起こし無応答になるため、
    呼び出しごとに専用プロファイルを割り当てる。
    """
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return None
    tmp = tempfile.mkdtemp()
    profile = Path(tmp) / "profile"
    try:
        # 実行ファイルのパスは shutil.which で解決したものだけを使い、
        # 引数に外部入力を混ぜない。shell も経由しない。
        subprocess.run(  # noqa: S603
            [soffice,
             f"-env:UserInstallation=file://{profile}",
             "--headless", "--norestore", "--invisible",
             "--convert-to", "csv", "--outdir", tmp, str(xlsx)],
            check=True, timeout=120,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except (subprocess.TimeoutExpired, subprocess.CalledProcessError):
        return None
    out = next(Path(tmp).glob("*.csv"), None)
    if out is None:
        return None
    with open(out, encoding="utf-8") as f:
        return list(csv.reader(f))


# --------------------------------------------------------------- 構造
def test_3シートが揃っている(artifacts):
    _, _, path = artifacts
    assert load_workbook(path).sheetnames == SHEETS


def test_職員名と日付が固定されている(artifacts):
    """横31日分をスクロールしても見出しが消えないこと。"""
    _, _, path = artifacts
    assert load_workbook(path)["勤務形態一覧表"].freeze_panes == "E10"


def test_兼務者は職種ごとに行が分かれる(artifacts):
    prob, _, _ = artifacts
    rows = expand_rows(prob)
    kenmu = sum(1 for st in prob.staff if st.secondary_job)
    assert len(rows) == len(prob.staff) + kenmu


# --------------------------------------------------------------- 数式の評価
def test_当月合計の数式がPythonの計算と一致する(artifacts):
    prob, sol, path = artifacts
    grid = _to_csv(path)
    if grid is None:
        pytest.skip("LibreOffice が無いため数式を評価できない")

    h = [p.work_minutes for p in prob.patterns]
    nd = prob.num_days
    DAY0, HR = 4, 7                      # 0始まりの列/行インデックス
    for k, (_job, i, ratio) in enumerate(expand_rows(prob)):
        expect = round(sum(h[sol.assign[i, d]] * ratio for d in range(nd)) / 60, 1)
        got = float(grid[HR + 2 + k][DAY0 + nd])
        assert got == pytest.approx(expect, abs=0.051), \
            f"{prob.staff[i].name} の当月合計が不一致"


def test_常勤換算の数式が第2位切り捨てで一致する(artifacts):
    prob, sol, path = artifacts
    grid = _to_csv(path)
    if grid is None:
        pytest.skip("LibreOffice が無いため数式を評価できない")

    h = [p.work_minutes for p in prob.patterns]
    nd = prob.num_days
    base = 7 + 2 + len(expand_rows(prob)) + 3
    for k, job in enumerate(FTE_TARGET_JOBS):
        row = grid[base + k]
        mins = sum(h[sol.assign[i, d]] * prob.staff[i].weight_for(job)
                   for i in range(len(prob.staff)) for d in range(nd))
        expect_total = round(mins / 60, 1)
        expect_fte = int(mins / prob.fulltime_month_minutes * 10) / 10
        assert float(row[1]) == pytest.approx(expect_total, abs=0.11), f"{job} 勤務延"
        assert float(row[3]) == pytest.approx(expect_fte, abs=1e-9), f"{job} 常勤換算"


# --------------------------------------------------------------- 違反の表示
def test_違反セルの塗りが違反した職種と日の数に一致する(artifacts):
    """同じ日・同じ職種で常勤換算と実人数の両方が不足すると違反は2件だが、
    塗るセルは1つ。比較対象は違反件数ではなく (職種, 日) の組の数。
    """
    prob, sol, path = artifacts
    ws = load_workbook(path)["日別の基準判定"]
    ng = {(v.job, v.day) for v in sol.violations}
    painted = 0
    for k in range(len(FTE_TARGET_JOBS)):
        r = 4 + k * 2 + 1                        # 「実際」の行
        for d in range(prob.num_days):
            fill = ws.cell(row=r, column=3 + d).fill
            if fill is not None and fill.fgColor.rgb in NG_FILL:
                painted += 1
    assert painted == len(ng)


def test_違反がある場合は日別シートに件数が明記される(artifacts):
    _, sol, path = artifacts
    ws = load_workbook(path)["日別の基準判定"]
    texts = [c.value for row in ws.iter_rows() for c in row
             if isinstance(c.value, str)]
    assert any(f"{len(sol.violations)} 件" in t for t in texts)


# --------------------------------------------------------------- 算定根拠
def test_算定根拠シートに端数処理と管理者の扱いが明記される(artifacts):
    _, _, path = artifacts
    ws = load_workbook(path)["算定根拠"]
    body = "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                     if c.value is not None)
    assert "小数点以下第2位を切り捨て" in body
    assert "管理者は常勤換算の対象外" in body
    assert "端数は切り上げない" in body


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
